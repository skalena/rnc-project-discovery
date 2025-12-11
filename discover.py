# -*- coding: utf-8 -*-
import os
import re
import sys
from datetime import datetime
from dataclasses import dataclass
from typing import Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import javalang
    HAS_JAVALANG = True
except ImportError:
    HAS_JAVALANG = False

# --- Configurações ---
# O caminho do projeto será solicitado via argumento de linha de comando.
JAVA_FILE_EXT = '.java'
JSF_FILE_EXT = ['.xhtml', '.jsf']
OUTPUT_FOLDER = 'output'

# Código de erro amigável quando a pasta do projeto não é fornecida
ERROR_CODE_MISSING_PROJECT = 2

# --- Estrutura de Dados para Análise de Regras de Negócios ---
@dataclass
class BusinessRuleMetrics:
    """Métricas de regras de negócio para uma classe"""
    class_name: str
    file_path: str
    controller_type: str
    public_methods: int = 0
    business_methods: int = 0
    business_method_names: List[str] = None
    
    def __post_init__(self):
        if self.business_method_names is None:
            self.business_method_names = []

# --- Padrões Comuns de Anotações/Arquivos ---
ENTITY_PATTERNS = [
    r'@Entity\b',
    r'@Table\b',
]

BUSINESS_PATTERNS = [
    r'@Named\b',
    r'@Controller\b',
    r'@Service\b',
    r'@RestController\b',
    r'@ManagedBean\b',
    r'extends.*Controller\b',
]

# --- Variáveis de Contagem ---
entity_classes = {}
jsf_pages = []
business_components = {}
db_info_placeholder = "Nenhuma informação de DB capturada de forma automática neste script."
analysis_log = ""
PROJECT_PATH = ""  # Será definido na execução
business_rules_metrics = {}  # Armazenar métricas de regras de negócio (será populado por analyze_business_rules)

def analyze_java_file(filepath):
    """Analisa um arquivo .java para Entidades e Componentes de Negócio."""
    global analysis_log
    content = ""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception as e:
        analysis_log += f"ERRO ao ler {filepath}: {e}\n"
        return

    found_entity_patterns = [p for p in ENTITY_PATTERNS if re.search(p, content)]
    if found_entity_patterns:
        entity_classes[filepath] = found_entity_patterns

    if not found_entity_patterns:
        found_business_patterns = [p for p in BUSINESS_PATTERNS if re.search(p, content)]
        if found_business_patterns:
            business_components[filepath] = found_business_patterns

def analyze_jsf_file(filepath):
    """Adiciona a página JSF à lista de encontrados."""
    jsf_pages.append(filepath)

def capture_database_info(project_root):
    """Tenta capturar informações básicas do DB."""
    global db_info_placeholder
    db_info_list = []
    config_files = ['.properties', '.xml', '.yml', '.yaml']

    for root, _, files in os.walk(project_root):
        for file in files:
            if any(file.endswith(ext) for ext in config_files):
                filepath = os.path.join(root, file)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        content = f.read()
                        info = []
                        if re.search(r'jdbc[:/]?.*url|jdbc:|spring.datasource.url', content, re.IGNORECASE):
                            info.append("Possível URL JDBC")
                        if re.search(r'hibernate\.dialect', content, re.IGNORECASE):
                            info.append("Dialeto Hibernate/JPA")
                        if re.search(r'spring\.datasource', content, re.IGNORECASE):
                            info.append("Configuração Spring Datasource")
                        if info:
                            db_info_list.append(f"- **{os.path.basename(file)}** ({', '.join(info)}) em `{filepath}`")
                except Exception:
                    pass

    if db_info_list:
        db_info_placeholder = "### Arquivos de Configuração de Banco de Dados Encontrados\n" + "\n".join(db_info_list)
    else:
        db_info_placeholder = "Não foram encontrados arquivos de configuração de DB comuns (.properties, .xml, .yml)."

def run_analysis(project_root):
    """Percorre a pasta do projeto e chama as funções de análise."""
    global analysis_log, business_rules_metrics

    analysis_log += f"Iniciando análise do projeto em: {project_root}\n"
    analysis_log += f"Hora de início: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"

    capture_database_info(project_root)

    for root, _, files in os.walk(project_root):
        for file in files:
            filepath = os.path.join(root, file)
            if file.endswith(JAVA_FILE_EXT):
                analyze_java_file(filepath)
            elif any(file.endswith(ext) for ext in JSF_FILE_EXT):
                analyze_jsf_file(filepath)

    analysis_log += "Análise de Arquivos Concluída.\n"
    
    # Realizar análise de regras de negócio com AST
    if HAS_JAVALANG:
        analysis_log += "Iniciando análise de regras de negócio (AST)...\n"
        business_rules_metrics = analyze_business_rules(project_root)
        if business_rules_metrics:
            analysis_log += f"  - Classes analisadas: {business_rules_metrics.get('total_classes', 0)}\n"
            analysis_log += f"  - Controllers encontrados: {business_rules_metrics.get('total_controllers', 0)}\n"
            analysis_log += f"  - Services encontrados: {business_rules_metrics.get('total_services', 0)}\n"
            analysis_log += f"  - Métodos com regras de negócio: {business_rules_metrics.get('total_business_methods', 0)}\n"
            if business_rules_metrics.get('total_controllers', 0) > 0:
                analysis_log += f"  - Média de métodos por Controller: {business_rules_metrics.get('avg_business_methods_per_controller', 0):.2f}\n"
        analysis_log += "Análise de regras de negócio concluída.\n"
    else:
        analysis_log += "⚠️  javalang não está disponível. Pulando análise de regras de negócio.\n"

def has_business_logic_in_method(method_node) -> bool:
    """
    Verifica se um método contém lógica de negócio analisando a árvore AST.
    Procura por statements de controle, operações de estado, etc.
    """
    if not hasattr(method_node, 'body') or not method_node.body:
        return False
    
    # Ignorar métodos muito simples (getters/setters)
    if method_node.name.startswith('get') or method_node.name.startswith('set') or method_node.name.startswith('is'):
        return False
    
    # Contar statements significativos na AST
    try:
        # Iterar pelos statements no corpo do método
        var_declarations = 0
        return_statements = 0
        body_size = len(str(method_node.body))
        
        for statement in method_node.body:
            # Verificar diferentes tipos de statements como indicadores de lógica de negócio
            
            # 1. Statements de controle = definitivamente lógica de negócio
            if isinstance(statement, (javalang.tree.IfStatement, javalang.tree.WhileStatement, 
                                    javalang.tree.ForStatement, javalang.tree.DoStatement,
                                    javalang.tree.SwitchStatement, javalang.tree.TryStatement)):
                return True
            
            # 2. Exceções = lógica de negócio
            if isinstance(statement, javalang.tree.ThrowStatement):
                return True
            
            # 3. Variáveis locais múltiplas + retorno = provavelmente processando dados
            if isinstance(statement, javalang.tree.LocalVariableDeclaration):
                var_declarations += 1
            
            # 4. Return statement
            if isinstance(statement, javalang.tree.ReturnStatement):
                return_statements += 1
        
        # Heurística: Se tem várias variáveis locais + return + tamanho significativo = lógica de negócio
        has_multiple_vars = var_declarations >= 3
        has_return = return_statements > 0
        has_significant_size = body_size > 1000  # 1KB+
        
        # Se tem múltiplas variáveis OU tamanho grande, é lógica de negócio
        if (has_multiple_vars and has_return) or has_significant_size:
            return True
        
        # Se tem pelo menos 2 variáveis + retorno + código de tamanho médio
        if var_declarations >= 2 and has_return and body_size > 500:
            return True
        
        return False
        
    except Exception:
        # Se houver erro ao processar AST, tenta fallback com string
        return False


def is_business_rule_method(method_node) -> bool:
    """
    Detecta se um método contém lógica de regra de negócio.
    Tenta análise AST primeiro, depois fallback para regex.
    """
    if not hasattr(method_node, 'body') or method_node.body is None:
        return False
    
    # Ignorar getters/setters/isXxx
    if any(method_node.name.startswith(prefix) for prefix in ['get', 'set', 'is']):
        return False
    
    # Tentar análise AST se javalang está disponível
    if HAS_JAVALANG:
        try:
            if has_business_logic_in_method(method_node):
                return True
        except Exception:
            pass  # Fallback para método de string
    
    # Fallback: análise de string
    method_body = str(method_node.body)
    lines_of_code = len([line for line in method_body.split('\n') if line.strip() and not line.strip().startswith('//')])
    
    # Métodos muito curtos são provavelmente simples
    if lines_of_code < 2:
        return False
    
    # Padrões que indicam regras de negócio
    business_patterns = [
        r'\bif\b', r'\belse\b', r'\bswitch\b', r'\bcase\b',      # Condicionais
        r'\bfor\b', r'\bwhile\b', r'\bdo\b',                      # Loops
        r'query\(', r'execute\(', r'save\(', r'delete\(',         # Operações de BD
        r'\.add\(', r'\.remove\(', r'\.set\(', r'\.put\(',        # Modificações de estado
        r'throw\s+', r'catch\s*\(',                                # Exceções
        r'return\s+[^;]*[\+\-\*/%]',                               # Cálculos no retorno
        r'\.compareTo\(', r'\.equals\(', r'\.contains\(',         # Comparações
    ]
    
    for pattern in business_patterns:
        if re.search(pattern, method_body):
            return True
    
    return False


def analyze_java_file_ast(file_path: str) -> Optional[List[BusinessRuleMetrics]]:
    """
    Analisa um arquivo Java usando AST e extrai métricas de regras de negócio.
    Retorna lista de BusinessRuleMetrics para cada classe/interface encontrada.
    """
    if not HAS_JAVALANG:
        return None
    
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        if not content.strip():
            return None
        
        tree = javalang.parse.parse(content)
        metrics_list = []
        
        # Iterar sobre tipos (classes, interfaces) declarados no arquivo
        for _, type_decl in tree.filter(javalang.tree.TypeDeclaration):
            class_name = type_decl.name
            
            # Contar métodos e identificar métodos com regras de negócio
            public_methods = 0
            business_methods = 0
            business_method_names = []
            
            for method in type_decl.methods:
                # Contar apenas métodos públicos
                if 'public' in method.modifiers:
                    public_methods += 1
                    
                    # Verificar se o método contém lógica de regra de negócio
                    if is_business_rule_method(method):
                        business_methods += 1
                        business_method_names.append(method.name)
            
            # Determinar tipo (Controller, Service, Repository, etc.)
            class_type = "Class"
            if hasattr(type_decl, 'name'):
                if 'Controller' in class_name:
                    class_type = "Controller"
                elif 'Service' in class_name:
                    class_type = "Service"
                elif 'Repository' in class_name:
                    class_type = "Repository"
                elif 'Impl' in class_name:
                    class_type = "Implementation"
            
            if public_methods > 0:  # Apenas incluir classes com métodos públicos
                metrics = BusinessRuleMetrics(
                    class_name=class_name,
                    file_path=file_path,
                    controller_type=class_type,
                    public_methods=public_methods,
                    business_methods=business_methods,
                    business_method_names=business_method_names
                )
                metrics_list.append(metrics)
        
        return metrics_list if metrics_list else None
    
    except Exception as e:
        # Silenciosamente falhar ao parsear - arquivo pode não ser Java válido
        return None


def analyze_business_rules(project_path: str) -> Dict[str, any]:
    """
    Analisa o projeto inteiro para regras de negócio e retorna métricas agregadas.
    """
    all_metrics = []
    controllers = []
    services = []
    
    # Procurar por arquivos Java
    for root, dirs, files in os.walk(project_path):
        for file in files:
            if file.endswith('.java'):
                file_path = os.path.join(root, file)
                metrics = analyze_java_file_ast(file_path)
                
                if metrics:
                    for metric in metrics:
                        all_metrics.append(metric)
                        if 'Controller' in metric.controller_type:
                            controllers.append(metric)
                        elif 'Service' in metric.controller_type:
                            services.append(metric)
    
    # Calcular estatísticas
    avg_business_methods_per_controller = 0
    avg_business_methods_per_service = 0
    total_business_methods = 0
    
    if controllers:
        total_business_methods_controllers = sum(m.business_methods for m in controllers)
        avg_business_methods_per_controller = total_business_methods_controllers / len(controllers)
    
    if services:
        total_business_methods_services = sum(m.business_methods for m in services)
        avg_business_methods_per_service = total_business_methods_services / len(services)
    
    if all_metrics:
        total_business_methods = sum(m.business_methods for m in all_metrics)
    
    return {
        'all_metrics': all_metrics,
        'controllers': controllers,
        'services': services,
        'total_classes': len(all_metrics),
        'total_controllers': len(controllers),
        'total_services': len(services),
        'avg_business_methods_per_controller': avg_business_methods_per_controller,
        'avg_business_methods_per_service': avg_business_methods_per_service,
        'total_business_methods': total_business_methods,
    }


def create_output_folder(project_path):
    """Cria a pasta de saída dentro do projeto se não existir."""
    output_path = os.path.join(project_path, OUTPUT_FOLDER)
    if not os.path.exists(output_path):
        os.makedirs(output_path)
        print(f"📁 Pasta '{output_path}' criada com sucesso.")
    return output_path

def generate_excel_report(folder_name, output_path):
    """Gera um relatório em formato Excel (.xlsx)."""
    if not HAS_OPENPYXL:
        print("⚠️  openpyxl não está instalado. Pulando geração de Excel.")
        return None

    try:
        wb = Workbook()
        wb.remove(wb.active)  # Remove a sheet padrão
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="FFFFFF")
        title_fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Sheet 1: Summary
        ws_summary = wb.create_sheet("Summary", 0)
        ws_summary['A1'] = "RNC Project Discovery - Analysis Report"
        ws_summary['A1'].font = title_font
        ws_summary['A1'].fill = title_fill
        ws_summary.merge_cells('A1:B1')
        
        ws_summary['A3'] = "Project Name:"
        ws_summary['B3'] = folder_name
        ws_summary['A4'] = "Analysis Date:"
        ws_summary['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws_summary['A5'] = "Project Path:"
        ws_summary['B5'] = PROJECT_PATH
        
        ws_summary['A7'] = "Metric"
        ws_summary['B7'] = "Count"
        for cell in ['A7', 'B7']:
            ws_summary[cell].font = header_font
            ws_summary[cell].fill = header_fill
            ws_summary[cell].alignment = center_align
        
        ws_summary['A8'] = "Entity Classes"
        ws_summary['B8'] = len(entity_classes)
        ws_summary['A9'] = "Business Components"
        ws_summary['B9'] = len(business_components)
        ws_summary['A10'] = "JSF Pages"
        ws_summary['B10'] = len(jsf_pages)
        
        for row in range(8, 11):
            ws_summary[f'A{row}'].alignment = left_align
            ws_summary[f'B{row}'].alignment = center_align
            ws_summary[f'B{row}'].font = Font(bold=True, size=11)

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 30

        # Sheet 2: Entity Classes
        ws_entities = wb.create_sheet("Entity Classes", 1)
        ws_entities['A1'] = "Entity Classes"
        ws_entities['A1'].font = title_font
        ws_entities['A1'].fill = title_fill
        ws_entities.merge_cells('A1:C1')
        
        ws_entities['A3'] = "File Path"
        ws_entities['B3'] = "Relative Path"
        ws_entities['C3'] = "Patterns Found"
        for cell in ['A3', 'B3', 'C3']:
            ws_entities[cell].font = header_font
            ws_entities[cell].fill = header_fill
            ws_entities[cell].alignment = center_align
            ws_entities[cell].border = border

        row = 4
        for filepath, patterns in entity_classes.items():
            relative_path = os.path.relpath(filepath, PROJECT_PATH)
            patterns_str = ', '.join(patterns)
            ws_entities[f'A{row}'] = filepath
            ws_entities[f'B{row}'] = relative_path
            ws_entities[f'C{row}'] = patterns_str
            for col in ['A', 'B', 'C']:
                ws_entities[f'{col}{row}'].alignment = left_align
                ws_entities[f'{col}{row}'].border = border
            row += 1

        ws_entities.column_dimensions['A'].width = 50
        ws_entities.column_dimensions['B'].width = 40
        ws_entities.column_dimensions['C'].width = 35

        # Sheet 3: Business Components
        ws_business = wb.create_sheet("Business Components", 2)
        ws_business['A1'] = "Business Components"
        ws_business['A1'].font = title_font
        ws_business['A1'].fill = title_fill
        ws_business.merge_cells('A1:C1')
        
        ws_business['A3'] = "File Path"
        ws_business['B3'] = "Relative Path"
        ws_business['C3'] = "Patterns Found"
        for cell in ['A3', 'B3', 'C3']:
            ws_business[cell].font = header_font
            ws_business[cell].fill = header_fill
            ws_business[cell].alignment = center_align
            ws_business[cell].border = border

        row = 4
        for filepath, patterns in business_components.items():
            relative_path = os.path.relpath(filepath, PROJECT_PATH)
            patterns_str = ', '.join(patterns)
            ws_business[f'A{row}'] = filepath
            ws_business[f'B{row}'] = relative_path
            ws_business[f'C{row}'] = patterns_str
            for col in ['A', 'B', 'C']:
                ws_business[f'{col}{row}'].alignment = left_align
                ws_business[f'{col}{row}'].border = border
            row += 1

        ws_business.column_dimensions['A'].width = 50
        ws_business.column_dimensions['B'].width = 40
        ws_business.column_dimensions['C'].width = 35

        # Sheet 4: JSF Pages
        ws_jsf = wb.create_sheet("JSF Pages", 3)
        ws_jsf['A1'] = "JSF Pages"
        ws_jsf['A1'].font = title_font
        ws_jsf['A1'].fill = title_fill
        ws_jsf.merge_cells('A1:B1')
        
        ws_jsf['A3'] = "File Path"
        ws_jsf['B3'] = "Relative Path"
        for cell in ['A3', 'B3']:
            ws_jsf[cell].font = header_font
            ws_jsf[cell].fill = header_fill
            ws_jsf[cell].alignment = center_align
            ws_jsf[cell].border = border

        row = 4
        for filepath in jsf_pages:
            relative_path = os.path.relpath(filepath, PROJECT_PATH)
            ws_jsf[f'A{row}'] = filepath
            ws_jsf[f'B{row}'] = relative_path
            for col in ['A', 'B']:
                ws_jsf[f'{col}{row}'].alignment = left_align
                ws_jsf[f'{col}{row}'].border = border
            row += 1

        ws_jsf.column_dimensions['A'].width = 50
        ws_jsf.column_dimensions['B'].width = 40

        # Sheet 5: Business Rules Analysis
        ws_rules = wb.create_sheet("Business Rules Analysis", 4)
        ws_rules['A1'] = "Business Rules Analysis"
        ws_rules['A1'].font = title_font
        ws_rules['A1'].fill = title_fill
        ws_rules.merge_cells('A1:F1')
        
        if business_rules_metrics and HAS_JAVALANG:
            # Headers
            headers = ["Class Name", "File", "Type", "Public Methods", "Business Rule Methods", "Business Method Names"]
            for col, header in enumerate(headers, start=1):
                cell = ws_rules.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            
            # Dados das classes
            row = 4
            all_metrics = business_rules_metrics.get('all_metrics', [])
            for metric in all_metrics:
                ws_rules[f'A{row}'] = metric.class_name
                ws_rules[f'B{row}'] = os.path.relpath(metric.file_path, PROJECT_PATH)
                ws_rules[f'C{row}'] = metric.controller_type
                ws_rules[f'D{row}'] = metric.public_methods
                ws_rules[f'E{row}'] = metric.business_methods
                ws_rules[f'F{row}'] = ', '.join(metric.business_method_names) if metric.business_method_names else ""
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws_rules[f'{col}{row}'].alignment = left_align
                    ws_rules[f'{col}{row}'].border = border
                row += 1
            
            # Resumo de estatísticas
            summary_row = row + 2
            ws_rules[f'A{summary_row}'] = "Summary Statistics"
            ws_rules[f'A{summary_row}'].font = Font(bold=True, size=11)
            
            summary_row += 1
            ws_rules[f'A{summary_row}'] = "Total Classes Analyzed:"
            ws_rules[f'B{summary_row}'] = business_rules_metrics.get('total_classes', 0)
            
            summary_row += 1
            ws_rules[f'A{summary_row}'] = "Total Controllers:"
            ws_rules[f'B{summary_row}'] = business_rules_metrics.get('total_controllers', 0)
            
            summary_row += 1
            ws_rules[f'A{summary_row}'] = "Total Services:"
            ws_rules[f'B{summary_row}'] = business_rules_metrics.get('total_services', 0)
            
            summary_row += 1
            ws_rules[f'A{summary_row}'] = "Total Business Rule Methods:"
            ws_rules[f'B{summary_row}'] = business_rules_metrics.get('total_business_methods', 0)
            
            summary_row += 1
            ws_rules[f'A{summary_row}'] = "Avg Business Methods per Controller:"
            ws_rules[f'B{summary_row}'] = f"{business_rules_metrics.get('avg_business_methods_per_controller', 0):.2f}"
            
            summary_row += 1
            ws_rules[f'A{summary_row}'] = "Avg Business Methods per Service:"
            ws_rules[f'B{summary_row}'] = f"{business_rules_metrics.get('avg_business_methods_per_service', 0):.2f}"
        else:
            ws_rules['A3'] = "Business rules analysis not available (javalang not installed)"
        
        ws_rules.column_dimensions['A'].width = 25
        ws_rules.column_dimensions['B'].width = 40
        ws_rules.column_dimensions['C'].width = 15
        ws_rules.column_dimensions['D'].width = 15
        ws_rules.column_dimensions['E'].width = 20
        ws_rules.column_dimensions['F'].width = 35

        # Sheet 6: Analysis Log
        ws_log = wb.create_sheet("Analysis Log", 5)
        ws_log['A1'] = "Analysis Log"
        ws_log['A1'].font = title_font
        ws_log['A1'].fill = title_fill
        ws_log.merge_cells('A1:A1')
        
        ws_log['A3'] = analysis_log
        ws_log['A3'].alignment = left_align
        ws_log.column_dimensions['A'].width = 80

        excel_filename = os.path.join(output_path, f"rnc-{folder_name}.xlsx")
        wb.save(excel_filename)
        return excel_filename

    except Exception as e:
        print(f"❌ Erro ao gerar relatório Excel: {e}")
        return None

def generate_markdown_report(output_path):
    """Gera o relatório final em formato Markdown."""
    folder_name = os.path.basename(os.path.abspath(PROJECT_PATH))
    report_filename = os.path.join(output_path, f"rnc-{folder_name}.md")

    report = f"# Relatório de Análise Estática do Projeto: `{folder_name}`\n\n"
    report += f"**Caminho do Projeto:** `{PROJECT_PATH}`\n"
    report += f"**Data da Análise:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
    report += "---\n\n"

    report += "## 1. Classes de Entidades/Objetos de Persistência\n\n"
    report += f"**Total de Classes Encontradas:** **{len(entity_classes)}**\n\n"
    report += "As classes foram identificadas pela presença de anotações JPA/Hibernate comuns (`@Entity`, `@Table`).\n\n"
    report += "```\n"
    for path, patterns in entity_classes.items():
        relative_path = os.path.relpath(path, PROJECT_PATH)
        patterns_str = ', '.join(p for p in patterns)
        report += f"* {relative_path} (Padrão: {patterns_str})\n"
    report += "```\n\n"

    report += "## 2. Classes de Componentes de Negócio/Controladoras/Backing Beans\n\n"
    report += f"**Total de Classes Encontradas:** **{len(business_components)}**\n\n"
    report += "As classes foram identificadas por anotações comuns de injeção/gerenciamento (`@Named`, `@Controller`, etc.).\n\n"
    report += "```\n"
    for path, patterns in business_components.items():
        relative_path = os.path.relpath(path, PROJECT_PATH)
        patterns_str = ', '.join(p for p in patterns)
        report += f"* {relative_path} (Padrão: {patterns_str})\n"
    report += "```\n\n"

    report += "## 3. Páginas JSF (XHTML) Encontradas\n\n"
    report += f"**Total de Páginas Encontradas:** **{len(jsf_pages)}**\n\n"
    report += "(A ligação exata entre a página JSF e a Entidade/Backing Bean é inferida por análise de código.)\n\n"
    report += "```\n"
    for path in jsf_pages:
        relative_path = os.path.relpath(path, PROJECT_PATH)
        report += f"* {relative_path}\n"
    report += "```\n\n"

    report += "## 4. Informações do Banco de Dados (Configurações)\n\n"
    report += "Análise simples de arquivos de configuração comuns:\n\n"
    report += db_info_placeholder
    report += "\n\n"

    report += "## 5. Análise de Regras de Negócio\n\n"
    if business_rules_metrics and HAS_JAVALANG:
        report += f"**Total de Classes Analisadas:** {business_rules_metrics.get('total_classes', 0)}\n\n"
        report += f"**Controllers Encontrados:** {business_rules_metrics.get('total_controllers', 0)}\n\n"
        report += f"**Services Encontrados:** {business_rules_metrics.get('total_services', 0)}\n\n"
        report += f"**Métodos com Regras de Negócio:** {business_rules_metrics.get('total_business_methods', 0)}\n\n"
        
        avg_per_controller = business_rules_metrics.get('avg_business_methods_per_controller', 0)
        report += f"**Número Médio de Métodos com Regras de Negócio por Controller:** `{avg_per_controller:.2f}`\n\n"
        
        avg_per_service = business_rules_metrics.get('avg_business_methods_per_service', 0)
        report += f"**Número Médio de Métodos com Regras de Negócio por Service:** `{avg_per_service:.2f}`\n\n"
        
        # Detalhar controllers com regras de negócio
        controllers = business_rules_metrics.get('controllers', [])
        if controllers:
            report += "### Controllers com Regras de Negócio\n\n"
            for controller in controllers:
                rel_path = os.path.relpath(controller.file_path, PROJECT_PATH)
                report += f"- **{controller.class_name}** ({rel_path})\n"
                report += f"  - Métodos públicos: {controller.public_methods}\n"
                report += f"  - Métodos com regras: {controller.business_methods}\n"
                if controller.business_method_names:
                    report += f"  - Métodos: {', '.join(controller.business_method_names)}\n"
                report += "\n"
        
        # Detalhar services com regras de negócio
        services = business_rules_metrics.get('services', [])
        if services:
            report += "### Services com Regras de Negócio\n\n"
            for service in services:
                rel_path = os.path.relpath(service.file_path, PROJECT_PATH)
                report += f"- **{service.class_name}** ({rel_path})\n"
                report += f"  - Métodos públicos: {service.public_methods}\n"
                report += f"  - Métodos com regras: {service.business_methods}\n"
                if service.business_method_names:
                    report += f"  - Métodos: {', '.join(service.business_method_names)}\n"
                report += "\n"
    else:
        report += "⚠️ Análise de regras de negócio não disponível (javalang não instalado).\n\n"

    report += "## 6. Log de Execução\n\n"
    report += "```\n"
    report += analysis_log
    report += "```\n"

    return report, report_filename

def generate_html_report(output_path):
    """Gera um relatório HTML profissional com CSS incorporado."""
    folder_name = os.path.basename(os.path.abspath(PROJECT_PATH))
    html_filename = os.path.join(output_path, f"rnc-{folder_name}.html")
    
    # Escape HTML entities
    def escape_html(text):
        return str(text).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;').replace("'", '&#39;')
    
    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>RNC Project Discovery - {escape_html(folder_name)}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
            min-height: 100vh;
            padding: 20px;
        }}
        
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.15);
            overflow: hidden;
        }}
        
        header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 40px 20px;
            text-align: center;
        }}
        
        header h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
            text-shadow: 0 2px 4px rgba(0,0,0,0.2);
        }}
        
        .meta-info {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 15px;
            margin-top: 20px;
            font-size: 0.95em;
            opacity: 0.95;
        }}
        
        .meta-info div {{
            background: rgba(255,255,255,0.1);
            padding: 10px 15px;
            border-radius: 6px;
            backdrop-filter: blur(10px);
        }}
        
        .meta-info strong {{
            display: block;
            margin-bottom: 5px;
            font-size: 0.85em;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        
        main {{
            padding: 40px;
        }}
        
        section {{
            margin-bottom: 50px;
        }}
        
        h2 {{
            font-size: 1.8em;
            color: #667eea;
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 3px solid #667eea;
            display: flex;
            align-items: center;
            gap: 10px;
        }}
        
        h3 {{
            font-size: 1.3em;
            color: #764ba2;
            margin: 25px 0 15px 0;
            margin-top: 30px;
        }}
        
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        
        .stat-card {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 25px;
            border-radius: 10px;
            text-align: center;
            box-shadow: 0 4px 15px rgba(102,126,234,0.3);
            transition: transform 0.3s ease;
        }}
        
        .stat-card:hover {{
            transform: translateY(-5px);
        }}
        
        .stat-card .number {{
            font-size: 2.5em;
            font-weight: bold;
            margin-bottom: 5px;
        }}
        
        .stat-card .label {{
            font-size: 0.9em;
            opacity: 0.9;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        
        thead {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }}
        
        th {{
            padding: 15px;
            text-align: left;
            font-weight: 600;
            text-transform: uppercase;
            font-size: 0.85em;
            letter-spacing: 0.5px;
        }}
        
        td {{
            padding: 12px 15px;
            border-bottom: 1px solid #eee;
        }}
        
        tbody tr:hover {{
            background: #f8f9ff;
        }}
        
        tbody tr:last-child td {{
            border-bottom: none;
        }}
        
        .file-list {{
            background: #f8f9ff;
            border-left: 4px solid #667eea;
            padding: 20px;
            border-radius: 6px;
            margin-bottom: 20px;
        }}
        
        .file-list ul {{
            list-style: none;
            margin-left: 0;
        }}
        
        .file-list li {{
            padding: 8px 0;
            border-bottom: 1px solid #ddd;
        }}
        
        .file-list li:last-child {{
            border-bottom: none;
        }}
        
        .file-list code {{
            background: white;
            padding: 3px 8px;
            border-radius: 4px;
            font-family: 'Courier New', monospace;
            font-size: 0.9em;
            color: #764ba2;
        }}
        
        .table-responsive {{
            overflow-x: auto;
            margin-bottom: 30px;
        }}
        
        .table-title {{
            font-size: 1.1em;
            color: #667eea;
            margin-bottom: 10px;
            font-weight: 600;
        }}
        
        .method-item {{
            background: #f8f9ff;
            padding: 15px;
            border-left: 4px solid #764ba2;
            margin-bottom: 15px;
            border-radius: 6px;
        }}
        
        .method-item h4 {{
            color: #764ba2;
            margin-bottom: 8px;
        }}
        
        .method-item .details {{
            font-size: 0.9em;
            color: #666;
            margin-top: 5px;
        }}
        
        .method-item .methods {{
            background: white;
            padding: 10px;
            border-radius: 4px;
            margin-top: 8px;
            font-family: 'Courier New', monospace;
            font-size: 0.85em;
            color: #667eea;
        }}
        
        .log-box {{
            background: #1e1e1e;
            color: #00ff00;
            padding: 20px;
            border-radius: 6px;
            font-family: 'Courier New', monospace;
            font-size: 0.9em;
            overflow-x: auto;
            white-space: pre-wrap;
            word-wrap: break-word;
            line-height: 1.5;
            border: 2px solid #333;
        }}
        
        footer {{
            background: #f5f7fa;
            padding: 30px 40px;
            text-align: center;
            color: #666;
            border-top: 1px solid #ddd;
            font-size: 0.9em;
        }}
        
        .icon {{
            display: inline-block;
            margin-right: 8px;
        }}
        
        .badge {{
            display: inline-block;
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 0.85em;
            font-weight: 600;
            margin-right: 8px;
            margin-top: 8px;
        }}
        
        .badge-success {{
            background: #d4edda;
            color: #155724;
        }}
        
        .badge-info {{
            background: #d1ecf1;
            color: #0c5460;
        }}
        
        .badge-warning {{
            background: #fff3cd;
            color: #856404;
        }}
        
        @media (max-width: 768px) {{
            header h1 {{
                font-size: 1.8em;
            }}
            
            main {{
                padding: 20px;
            }}
            
            .stats-grid {{
                grid-template-columns: 1fr;
            }}
            
            table {{
                font-size: 0.85em;
            }}
            
            th, td {{
                padding: 8px;
            }}
            
            .table-responsive {{
                overflow-x: auto;
                -webkit-overflow-scrolling: touch;
            }}
            
            table {{
                min-width: 500px;
            }}
            
            h2 {{
                font-size: 1.3em;
            }}
            
            h3 {{
                font-size: 1.1em;
            }}
        }}
        
        @media (max-width: 480px) {{
            header h1 {{
                font-size: 1.4em;
            }}
            
            main {{
                padding: 15px;
            }}
            
            table {{
                font-size: 0.75em;
            }}
            
            th, td {{
                padding: 6px;
            }}
            
            .stat-card {{
                padding: 15px;
            }}
            
            .stat-card .number {{
                font-size: 2em;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>📊 RNC Project Discovery</h1>
            <p>Análise Estática de Projeto Java e JSF</p>
            <div class="meta-info">
                <div>
                    <strong>Projeto</strong>
                    {escape_html(folder_name)}
                </div>
                <div>
                    <strong>Data Análise</strong>
                    {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
                </div>
                <div>
                    <strong>Caminho</strong>
                    {escape_html(PROJECT_PATH)}
                </div>
            </div>
        </header>
        
        <main>
            <!-- SEÇÃO 1: RESUMO EXECUTIVO -->
            <section>
                <h2>📈 Resumo Executivo</h2>
                <div class="stats-grid">
                    <div class="stat-card">
                        <div class="number">{len(entity_classes)}</div>
                        <div class="label">Classes de Entidades</div>
                    </div>
                    <div class="stat-card">
                        <div class="number">{len(business_components)}</div>
                        <div class="label">Componentes de Negócio</div>
                    </div>
                    <div class="stat-card">
                        <div class="number">{len(jsf_pages)}</div>
                        <div class="label">Páginas JSF</div>
                    </div>
                    <div class="stat-card">
                        <div class="number">{business_rules_metrics.get('total_business_methods', 0)}</div>
                        <div class="label">Métodos com Regras</div>
                    </div>
                </div>
            </section>
            
            <!-- SEÇÃO 2: CLASSES DE ENTIDADES -->
            <section>
                <h2>🗄️ Classes de Entidades/Objetos de Persistência</h2>
                <p><strong>Total encontrado:</strong> <span class="badge badge-info">{len(entity_classes)}</span></p>
                <p>Classes identificadas pela presença de anotações JPA/Hibernate comuns (<code>@Entity</code>, <code>@Table</code>).</p>
                <div class="table-responsive">
                    <div class="table-title">Detalhes das Classes de Entidades</div>
                    <table>
                        <thead>
                            <tr>
                                <th>Caminho do Arquivo</th>
                                <th>Padrões Detectados</th>
                            </tr>
                        </thead>
                        <tbody>
"""
    
    # Adicionar entidades em tabela
    for filepath, patterns in entity_classes.items():
        relative_path = os.path.relpath(filepath, PROJECT_PATH)
        patterns_str = ', '.join(patterns) if patterns else 'N/A'
        html += f'                            <tr><td><code>{escape_html(relative_path)}</code></td><td>{escape_html(patterns_str)}</td></tr>\n'
    
    html += """                        </tbody>
                    </table>
                </div>
            </section>
            
            <!-- SEÇÃO 3: COMPONENTES DE NEGÓCIO -->
            <section>
                <h2>🔧 Classes de Componentes de Negócio/Controladoras</h2>
                <p><strong>Total encontrado:</strong> <span class="badge badge-success">{}</span></p>
                <p>Classes identificadas por anotações comuns de injeção/gerenciamento (<code>@Named</code>, <code>@Controller</code>, etc.).</p>
                <div class="table-responsive">
                    <div class="table-title">Detalhes dos Componentes de Negócio</div>
                    <table>
                        <thead>
                            <tr>
                                <th>Caminho do Arquivo</th>
                                <th>Anotações Detectadas</th>
                            </tr>
                        </thead>
                        <tbody>
""".format(len(business_components))
    
    # Adicionar componentes em tabela
    for filepath, patterns in business_components.items():
        relative_path = os.path.relpath(filepath, PROJECT_PATH)
        patterns_str = ', '.join(patterns) if patterns else 'N/A'
        html += f'                            <tr><td><code>{escape_html(relative_path)}</code></td><td>{escape_html(patterns_str)}</td></tr>\n'
    
    html += """                        </tbody>
                    </table>
                </div>
            </section>
            
            <!-- SEÇÃO 4: PÁGINAS JSF -->
            <section>
                <h2>🖼️ Páginas JSF (XHTML)</h2>
                <p><strong>Total encontrado:</strong> <span class="badge badge-warning">{}</span></p>
                <p>Arquivos de view utilizados em aplicações JSF.</p>
                <div class="table-responsive">
                    <div class="table-title">Detalhes das Páginas JSF</div>
                    <table>
                        <thead>
                            <tr>
                                <th>Caminho do Arquivo</th>
                                <th>Tipo</th>
                            </tr>
                        </thead>
                        <tbody>
""".format(len(jsf_pages))
    
    # Adicionar páginas JSF em tabela
    for filepath in jsf_pages:
        relative_path = os.path.relpath(filepath, PROJECT_PATH)
        file_type = os.path.splitext(filepath)[1]
        html += f'                            <tr><td><code>{escape_html(relative_path)}</code></td><td>{escape_html(file_type)}</td></tr>\n'
    
    html += """                        </tbody>
                    </table>
                </div>
            </section>
            
            <!-- SEÇÃO 5: ANÁLISE DE REGRAS DE NEGÓCIO -->
            <section>
                <h2>🧠 Análise de Regras de Negócio</h2>
"""
    
    if business_rules_metrics and HAS_JAVALANG:
        html += f"""                <div class="stats-grid">
                    <div class="stat-card">
                        <div class="number">{business_rules_metrics.get('total_classes', 0)}</div>
                        <div class="label">Classes Analisadas</div>
                    </div>
                    <div class="stat-card">
                        <div class="number">{business_rules_metrics.get('total_controllers', 0)}</div>
                        <div class="label">Controllers</div>
                    </div>
                    <div class="stat-card">
                        <div class="number">{business_rules_metrics.get('total_services', 0)}</div>
                        <div class="label">Services</div>
                    </div>
                    <div class="stat-card">
                        <div class="number">{business_rules_metrics.get('avg_business_methods_per_service', 0):.2f}</div>
                        <div class="label">Média por Service</div>
                    </div>
                </div>
"""
        
        # Controllers com regras
        controllers = business_rules_metrics.get('controllers', [])
        if controllers:
            html += """                <h3>Controllers com Regras de Negócio</h3>
                <div class="table-responsive">
                    <table>
                        <thead>
                            <tr>
                                <th>Classe</th>
                                <th>Arquivo</th>
                                <th>Métodos Públicos</th>
                                <th>Métodos com Regras</th>
                                <th>Nomes dos Métodos</th>
                            </tr>
                        </thead>
                        <tbody>
"""
            for controller in controllers:
                rel_path = os.path.relpath(controller.file_path, PROJECT_PATH)
                methods_str = ', '.join(controller.business_method_names) if controller.business_method_names else '-'
                html += f"""                            <tr>
                                <td><strong>{escape_html(controller.class_name)}</strong></td>
                                <td><code>{escape_html(rel_path)}</code></td>
                                <td>{controller.public_methods}</td>
                                <td><span class="badge badge-success">{controller.business_methods}</span></td>
                                <td>{escape_html(methods_str)}</td>
                            </tr>
"""
            html += """                        </tbody>
                    </table>
                </div>
"""
        
        # Services com regras
        services = business_rules_metrics.get('services', [])
        if services:
            html += """                <h3>Services com Regras de Negócio</h3>
                <div class="table-responsive">
                    <table>
                        <thead>
                            <tr>
                                <th>Classe</th>
                                <th>Arquivo</th>
                                <th>Métodos Públicos</th>
                                <th>Métodos com Regras</th>
                                <th>Nomes dos Métodos</th>
                            </tr>
                        </thead>
                        <tbody>
"""
            for service in services:
                rel_path = os.path.relpath(service.file_path, PROJECT_PATH)
                methods_str = ', '.join(service.business_method_names) if service.business_method_names else '-'
                html += f"""                            <tr>
                                <td><strong>{escape_html(service.class_name)}</strong></td>
                                <td><code>{escape_html(rel_path)}</code></td>
                                <td>{service.public_methods}</td>
                                <td><span class="badge badge-success">{service.business_methods}</span></td>
                                <td>{escape_html(methods_str)}</td>
                            </tr>
"""
            html += """                        </tbody>
                    </table>
                </div>
"""
    else:
        html += "                <p><em>⚠️ Análise de regras de negócio não disponível (javalang não instalado).</em></p>\n"
    
    # SEÇÃO 6: LOG
    html += f"""            </section>
            
            <!-- SEÇÃO 6: LOG DE EXECUÇÃO -->
            <section>
                <h2>📋 Log de Execução</h2>
                <div class="log-box">{escape_html(analysis_log)}</div>
            </section>
        </main>
        
        <footer>
            <p>📊 Relatório gerado automaticamente por RNC Project Discovery</p>
            <p style="margin-top: 10px; font-size: 0.85em; color: #999;">
                Versão 1.0 | {datetime.now().strftime('%d de %B de %Y às %H:%M:%S')}
            </p>
        </footer>
    </div>
</body>
</html>
"""
    
    try:
        with open(html_filename, 'w', encoding='utf-8') as f:
            f.write(html)
        return html_filename
    except Exception as e:
        print(f"❌ Erro ao gerar relatório HTML: {e}")
        return None

def save_and_display_report(report_content, md_filename, excel_filename=None):
    """Salva os relatórios em Markdown, Excel e HTML, e exibe o resumo no terminal."""
    try:
        with open(md_filename, 'w', encoding='utf-8') as f:
            f.write(report_content)
        print(f"✅ Relatório Markdown salvo em: {md_filename}")
    except Exception as e:
        print(f"❌ Erro ao salvar o arquivo Markdown {md_filename}: {e}")

    if excel_filename:
        print(f"✅ Relatório Excel salvo em: {excel_filename}")

    # Gerar HTML
    output_path = os.path.dirname(md_filename)
    html_filename = generate_html_report(output_path)
    if html_filename:
        print(f"✅ Relatório HTML salvo em: {html_filename}")

    print("\n" + "="*80)
    print("RESUMO DA ANÁLISE ESTATICA")
    print("="*80)
    print(f"Projeto Analisado: {os.path.basename(os.path.abspath(PROJECT_PATH))}")
    print(f"Total de Entidades: {len(entity_classes)}")
    print(f"Total de Componentes de Negócio/Controladoras: {len(business_components)}")
    print(f"Total de Páginas JSF: {len(jsf_pages)}")
    found_db = "Sim" if "Arquivos de Configuração de Banco de Dados Encontrados" in db_info_placeholder else "Não"
    print(f"Informações de DB Encontradas: {found_db}")
    print("="*80)
    print(f"📂 Arquivos de saída estão em: {os.path.dirname(md_filename)}/")
    print("="*80)

# --- Execução Principal ---
if __name__ == "__main__":

    # Requer que o caminho do projeto seja passado como argumento.
    # Se desejar entrada interativa, passe a flag --interactive; caso contrário o script retornará um erro amigável.
    if len(sys.argv) <= 1:
        # Mensagem de erro amigável em texto e JSON para automações
        usage = f"Uso: python {os.path.basename(__file__)} <caminho_para_pasta_projeto>\nExemplo: python {os.path.basename(__file__)} C:\\meu_projeto"
        message_text = (
            "\n🛑 Erro: Pasta do projeto não informada.\n\n"
            f"{usage}\n"
            f"Exemplo: python {os.path.basename(__file__)} C:\\meu_projeto\n\n"
            f"Código de erro: {ERROR_CODE_MISSING_PROJECT}\n"
        )
        print(message_text)
        sys.exit(ERROR_CODE_MISSING_PROJECT)

    PROJECT_PATH = sys.argv[1]
    PROJECT_PATH = os.path.abspath(PROJECT_PATH)

    if not os.path.isdir(PROJECT_PATH):
        print(f"\n🛑 ERRO: O caminho '{PROJECT_PATH}' não é um diretório válido.")
        sys.exit(3)

    # Criar pasta de output dentro do projeto
    output_folder_path = create_output_folder(PROJECT_PATH)

    # Executar análise
    run_analysis(PROJECT_PATH)

    # Gerar relatórios
    report_content, md_filename = generate_markdown_report(output_folder_path)
    excel_filename = generate_excel_report(os.path.basename(os.path.abspath(PROJECT_PATH)), output_folder_path)

    # Salvar e exibir resultados
    save_and_display_report(report_content, md_filename, excel_filename)